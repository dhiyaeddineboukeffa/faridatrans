"""
ETL Script: Parse BUS DE CONSTANTINE.xlsx, les arrets des bus.xlsx, tramway.constantine.xlsx → network_data.json
"""
import openpyxl
import json
import re
from math import radians, cos, sin, asin, sqrt

BUS_EXCEL_PATH = r"BUS DE CONSTANTINE.xlsx"
ARRETS_EXCEL_PATH = r"les arrets des bus.xlsx"
TRAMWAY_EXCEL_PATH = r"tramway.constantine.xlsx"
OUTPUT_PATH = r"backend/routing-service/network_data.json"
STATIONS_PATH = r"backend/routing-service/stations.json"

def parse_coords(raw):
    """Parse a string like '36.5663 , 6.7369' into (lat, lon)."""
    if not raw:
        return None, None
    parts = re.split(r'[,\s]+', str(raw).strip())
    parts = [p for p in parts if p]
    if len(parts) >= 2:
        try:
            return float(parts[0]), float(parts[1])
        except ValueError:
            return None, None
    return None, None

def slugify(name):
    """Create a simple stop ID from a name."""
    return str(name).strip().upper().replace(" ", "_").replace("'", "").replace("-", "_").replace("/", "_")

def calc_dist(lat1, lon1, lat2, lon2):
    R = 6371000  # Earth radius in meters
    lat1, lon1 = radians(lat1), radians(lon1)
    lat2, lon2 = radians(lat2), radians(lon2)
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    return 2 * R * asin(sqrt(a))

nodes = {}  # id -> {name, lat, lon, type}
edges = []  # {u, v, mode, route, duration, weight}

# ============================================================
# 1. Parse BUS DE CONSTANTINE (city-wide bus lines, start/end only)
# ============================================================
print("Parsing BUS DE CONSTANTINE.xlsx...")
try:
    wb_bus = openpyxl.load_workbook(BUS_EXCEL_PATH)
    ws_bus = wb_bus.active
    for row in ws_bus.iter_rows(min_row=2, values_only=True):
        if len(row) < 7:
            continue
        line_num, dep_name, dep_coords, arr_name, arr_coords = row[0], row[1], row[3], row[4], row[6]

        if not dep_name or not arr_name:
            continue

        dep_lat, dep_lon = parse_coords(dep_coords)
        arr_lat, arr_lon = parse_coords(arr_coords)

        if dep_lat is None or arr_lat is None:
            continue

        dep_id = slugify(dep_name)
        arr_id = slugify(arr_name)
        route_id = f"L{line_num}"

        if dep_id not in nodes:
            nodes[dep_id] = {"name": str(dep_name).strip(), "lat": dep_lat, "lon": dep_lon, "type": "bus"}
        if arr_id not in nodes:
            nodes[arr_id] = {"name": str(arr_name).strip(), "lat": arr_lat, "lon": arr_lon, "type": "bus"}

        dist_m = calc_dist(dep_lat, dep_lon, arr_lat, arr_lon)
        duration_sec = int(dist_m / (30 * 1000 / 3600))  # 30 km/h

        edges.append({"u": dep_id, "v": arr_id, "mode": "bus", "route": route_id, "duration": duration_sec, "weight": duration_sec})
        edges.append({"u": arr_id, "v": dep_id, "mode": "bus", "route": route_id, "duration": duration_sec, "weight": duration_sec})
except Exception as e:
    print(f"Error parsing BUS DE CONSTANTINE: {e}")

# ============================================================
# 2. Parse Tramway (bidirectional, already good)
# ============================================================
print("Parsing tramway.constantine.xlsx...")
try:
    wb_tram = openpyxl.load_workbook(TRAMWAY_EXCEL_PATH)
    ws_tram = wb_tram.active
    
    prev_stop_id = None
    prev_lat, prev_lon = None, None
    
    for row in ws_tram.iter_rows(min_row=2, values_only=True):
        if len(row) < 4:
            continue
        
        stop_name = row[1]
        coords_raw = row[3]
        
        if not stop_name or not coords_raw:
            continue
            
        lat, lon = parse_coords(coords_raw)
        if lat is None:
            continue
            
        stop_id = slugify(stop_name)
        if stop_id not in nodes:
            nodes[stop_id] = {"name": str(stop_name).strip(), "lat": lat, "lon": lon, "type": "tram"}
            
        if prev_stop_id is not None:
            dist_m = calc_dist(prev_lat, prev_lon, lat, lon)
            duration_sec = int(dist_m / (20 * 1000 / 3600)) # 20 km/h for tram
            edges.append({"u": prev_stop_id, "v": stop_id, "mode": "tram", "route": "T1", "duration": duration_sec, "weight": duration_sec})
            edges.append({"u": stop_id, "v": prev_stop_id, "mode": "tram", "route": "T1", "duration": duration_sec, "weight": duration_sec})
            
        prev_stop_id = stop_id
        prev_lat, prev_lon = lat, lon
except Exception as e:
    print(f"Error parsing tramway: {e}")

# ============================================================
# 3. Parse Bus Stops (les arrets des bus.xlsx)
#    - Dynamically detect name vs coordinate columns
#    - Make ALL edges bidirectional
# ============================================================
print("Parsing les arrets des bus.xlsx...")
try:
    wb_arrets = openpyxl.load_workbook(ARRETS_EXCEL_PATH)
    ws_arrets = wb_arrets.active

    cols = list(ws_arrets.iter_cols(values_only=True))
    
    def is_coord_col(col):
        """Check if a column contains coordinates by testing row 2."""
        if len(col) < 3 or col[2] is None:
            return False
        lat, lon = parse_coords(col[2])
        return lat is not None
    
    routes_parsed = 0
    i = 0
    while i < len(cols):
        col = cols[i]
        
        # Skip empty columns
        if all(c is None for c in col):
            i += 1
            continue
        
        # Determine if this column is a names column or coords column
        if is_coord_col(col):
            coord_col = col
            name_col = cols[i - 1] if i > 0 and not is_coord_col(cols[i - 1]) else None
            i += 1
        elif i + 1 < len(cols) and is_coord_col(cols[i + 1]):
            name_col = col
            coord_col = cols[i + 1]
            i += 2
        else:
            i += 1
            continue
        
        route_name = coord_col[1]
        if not route_name or str(route_name).strip() == "":
            continue
        
        route_name = str(route_name).strip()
        route_id = f"BUS_{slugify(route_name)}"
        
        # Collect all stops for this route direction
        route_stops = []
        
        max_len = max(len(name_col) if name_col else 0, len(coord_col))
        for idx in range(2, max_len):
            cell_val = coord_col[idx] if idx < len(coord_col) else None
            name_val = name_col[idx] if name_col and idx < len(name_col) else None
            
            if not cell_val:
                continue
            
            lat, lon = parse_coords(cell_val)
            if lat is None:
                continue
            
            stop_id = f"STOP_{lat:.5f}_{lon:.5f}"
            stop_name = str(name_val).strip() if name_val and str(name_val).strip() else f"{route_name} (Stop {idx - 1})"

            if stop_id not in nodes:
                nodes[stop_id] = {"name": stop_name, "lat": lat, "lon": lon, "type": "bus"}
            else:
                # Update name if we have a better one than the generic
                if "(Stop" not in stop_name and "(Stop" in nodes[stop_id]["name"]:
                    nodes[stop_id]["name"] = stop_name
            
            route_stops.append((stop_id, lat, lon))
        
        # Create BIDIRECTIONAL edges for this route
        for j in range(len(route_stops) - 1):
            sid_a, lat_a, lon_a = route_stops[j]
            sid_b, lat_b, lon_b = route_stops[j + 1]
            
            if sid_a == sid_b:
                continue
                
            dist_m = calc_dist(lat_a, lon_a, lat_b, lon_b)
            duration_sec = int(dist_m / (30 * 1000 / 3600))  # 30 km/h
            
            # Forward edge
            edges.append({"u": sid_a, "v": sid_b, "mode": "bus", "route": route_id, "duration": duration_sec, "weight": duration_sec})
            # Reverse edge (same route, opposite direction)
            edges.append({"u": sid_b, "v": sid_a, "mode": "bus", "route": route_id, "duration": duration_sec, "weight": duration_sec})
        
        routes_parsed += 1
    
    print(f"  Parsed {routes_parsed} route directions from arrets file.")
except Exception as e:
    print(f"Error parsing arrets: {e}")

# ============================================================
# 3. Walking edges (connect nearby stops from different lines)
# ============================================================
print("Calculating walking edges...")
stop_list = list(nodes.items())
walk_added = 0
for i in range(len(stop_list)):
    for j in range(i + 1, len(stop_list)):
        id_a, s_a = stop_list[i]
        id_b, s_b = stop_list[j]
        dist_m = calc_dist(s_a['lat'], s_a['lon'], s_b['lat'], s_b['lon'])
        if dist_m < 500:  # 500m walking radius
            walk_sec = int(dist_m / 1.4)  # 1.4 m/s walking
            edges.append({"u": id_a, "v": id_b, "mode": "walk", "route": None, "duration": walk_sec, "weight": walk_sec})
            edges.append({"u": id_b, "v": id_a, "mode": "walk", "route": None, "duration": walk_sec, "weight": walk_sec})
            walk_added += 1

# ============================================================
# 4. Write output
# ============================================================
output = {"nodes": nodes, "edges": edges}
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

# stations.json — ALL stops now (not just named ones)
station_list = [
    {"id": stop_id, **data}
    for stop_id, data in nodes.items()
]
with open(STATIONS_PATH, "w", encoding="utf-8") as f:
    json.dump(station_list, f, indent=2, ensure_ascii=False)

print(f"\n[SUCCESS] Done!")
print(f"   Stops (nodes):  {len(nodes)}")
print(f"   Stations:       {len(station_list)} (written to stations.json)")
print(f"   Route edges:    {len([e for e in edges if e['mode'] in ['bus', 'tram']])}")
print(f"   Walk edges:     {len([e for e in edges if e['mode'] == 'walk'])} (from {walk_added} pairs)")
print(f"   Output:         {OUTPUT_PATH}")
